"""Build the searchable index: PDFs + the profiles/accessories workbook -> SQLite FTS5.

Per PDF page we store the text, a full-page render, a thumbnail, every embedded
raster, and - importantly for these catalogues - rendered crops of the *vector*
drawings. Installux technical pages carry thousands of vector strokes and almost
no raster images, so without the crop pass there is nothing to show the user
when they ask for a cross-section.

Identical PDFs filed under several system folders (Galaxie 45TH currently holds
byte-for-byte copies of the Galaxie 32TH files) are indexed once and tagged with
every folder they appear in, so results are not duplicated.

Rendering and hashing run across all CPU cores; workers produce rows and files,
and the parent process is the only writer to SQLite.
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import sqlite3
from collections import deque
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

import fitz  # PyMuPDF
import numpy as np

import imgutil

BASE_DIR = Path(__file__).resolve().parent
PDF_DIR = BASE_DIR / "pdfs"
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "index.db"
PAGES_DIR = DATA_DIR / "pages"
FIGURES_DIR = DATA_DIR / "figures"
REF_IMAGES_DIR = PDF_DIR / "Images"

THUMB_DPI = 90
FULL_DPI = 150
FIGURE_DPI = 200
FIG_THUMB_W = 320
MAX_IMAGE_SIDE = 2200

# vector-drawing clustering
CELL_PT = 6.0            # occupancy grid resolution, in PDF points
DILATE = 2               # cells: merges strokes belonging to one figure
MIN_AREA_FRAC = 0.010    # ignore specks (rules, page furniture)
MAX_AREA_FRAC = 0.88     # ignore full-page backgrounds (the page render covers those)
MAX_FIGURES_PER_PAGE = 8
PAD_PT = 7.0

MAX_WORKERS = 8
REF_CHUNK = 250

# Installux part references: 102, 410031, L1502, 10230A, 11-140, 12AR05, 10X2-T
REF_RE = re.compile(r"\b[A-Z]{0,3}\d{1,3}[A-Z0-9]*(?:[-/.][A-Z0-9]+)*\b")
# TOC page ranges like "pages 50-51" must not be indexed as part 5051
_PAGE_RANGE_RE = re.compile(r"^\d{1,3}-\d{1,3}$")


def sanitize(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_") or "doc"


def norm_ref(ref: str) -> str:
    """Comparison key for a reference: upper case, separators removed."""
    return re.sub(r"[\s./\-]", "", str(ref).upper())


def worker_count(requested: int = 0) -> int:
    if requested and requested > 0:
        return min(requested, MAX_WORKERS)
    return max(1, min(MAX_WORKERS, (os.cpu_count() or 2)))


# --------------------------------------------------------------------------
# schema
# --------------------------------------------------------------------------
def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA journal_mode=WAL;
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY,
            sha1 TEXT NOT NULL,
            filename TEXT NOT NULL,
            title TEXT,
            system TEXT DEFAULT '',
            systems TEXT DEFAULT '[]',
            systems_text TEXT DEFAULT '',
            doc_kind TEXT DEFAULT 'document',
            num_pages INTEGER,
            indexed_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS pages (
            id INTEGER PRIMARY KEY,
            doc_id INTEGER NOT NULL REFERENCES documents(id),
            page_num INTEGER NOT NULL,
            text TEXT NOT NULL DEFAULT '',
            ocr_text TEXT NOT NULL DEFAULT '',
            refs TEXT NOT NULL DEFAULT '',
            summary TEXT DEFAULT '',
            vec TEXT
        );
        CREATE TABLE IF NOT EXISTS figures (
            id INTEGER PRIMARY KEY,
            doc_id INTEGER NOT NULL REFERENCES documents(id),
            page_num INTEGER NOT NULL,
            kind TEXT NOT NULL DEFAULT 'raster',
            filename TEXT NOT NULL,
            thumb TEXT,
            width INTEGER,
            height INTEGER,
            area REAL DEFAULT 0,
            x0 REAL DEFAULT 0, y0 REAL DEFAULT 0, x1 REAL DEFAULT 0, y1 REAL DEFAULT 0,
            phash TEXT, dhash TEXT, edge TEXT
        );
        CREATE TABLE IF NOT EXISTS parts (
            id INTEGER PRIMARY KEY,
            ref TEXT NOT NULL,
            ref_norm TEXT NOT NULL DEFAULT '',
            kind TEXT DEFAULT 'part',
            supplier_ref TEXT DEFAULT '',
            designation TEXT DEFAULT '',
            raw TEXT DEFAULT '',
            vec TEXT
        );
        CREATE TABLE IF NOT EXISTS ref_images (
            id INTEGER PRIMARY KEY,
            ref TEXT NOT NULL,
            ref_norm TEXT NOT NULL DEFAULT '',
            filename TEXT NOT NULL,
            phash TEXT, dhash TEXT, edge TEXT
        );
        CREATE TABLE IF NOT EXISTS ocr_boxes (
            id INTEGER PRIMARY KEY,
            doc_id INTEGER NOT NULL REFERENCES documents(id),
            page_num INTEGER NOT NULL,
            text TEXT NOT NULL,
            x REAL, y REAL, w REAL, h REAL
        );
        CREATE INDEX IF NOT EXISTS idx_ocr_boxes ON ocr_boxes(doc_id, page_num);
        CREATE INDEX IF NOT EXISTS idx_pages_doc ON pages(doc_id, page_num);
        CREATE INDEX IF NOT EXISTS idx_figures_doc ON figures(doc_id, page_num);
        CREATE INDEX IF NOT EXISTS idx_parts_ref ON parts(ref_norm);
        CREATE INDEX IF NOT EXISTS idx_ref_images_ref ON ref_images(ref_norm);
        """
    )
    # separate columns let the ranker weight a reference hit above prose
    conn.execute(
        "CREATE VIRTUAL TABLE IF NOT EXISTS pages_fts "
        "USING fts5(text, refs, title, ocr, tokenize='unicode61 remove_diacritics 2')"
    )
    conn.execute(
        "CREATE VIRTUAL TABLE IF NOT EXISTS parts_fts "
        "USING fts5(ref, designation, tokenize='unicode61 remove_diacritics 2')"
    )


# --------------------------------------------------------------------------
# document metadata
# --------------------------------------------------------------------------
SYSTEM_PATTERNS = [
    ("COMETE 70TH", re.compile(r"com[eè]te|70\s*th", re.I)),
    ("GALAXIE 45TH", re.compile(r"galaxie\s*45|45\s*th", re.I)),
    ("GALAXIE 32TH", re.compile(r"galaxie\s*32|32\s*th", re.I)),
]


def detect_system(folder: str, filename: str) -> str:
    for name, rx in SYSTEM_PATTERNS:
        if rx.search(folder):
            return name
    for name, rx in SYSTEM_PATTERNS:
        if rx.search(filename):
            return name
    return folder.upper() or "GENERAL"


def detect_kind(filename: str) -> str:
    f = filename.lower()
    if re.match(r"^d\d{3,5}", f):
        return "fabrication drawing"
    if "doutil" in f or "tool" in f:
        return "technical catalogue"
    return "product brochure"


def sha1_of(path: Path) -> str:
    h = hashlib.sha1()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# --------------------------------------------------------------------------
# vector-drawing extraction
# --------------------------------------------------------------------------
def _occupancy(page) -> tuple[np.ndarray, fitz.Rect] | None:
    """Mark which grid cells any vector stroke touches."""
    pr = page.rect
    if pr.width <= 0 or pr.height <= 0:
        return None
    try:
        drawings = page.get_drawings()
    except Exception:
        return None
    if not drawings:
        return None

    nx = max(1, int(pr.width / CELL_PT) + 1)
    ny = max(1, int(pr.height / CELL_PT) + 1)
    page_area = pr.width * pr.height
    grid = np.zeros((ny, nx), dtype=bool)
    for d in drawings:
        r = d.get("rect")
        if r is None or r.is_empty or r.is_infinite:
            continue
        r = r & pr
        if r.is_empty:
            continue
        # a stroke spanning most of the sheet is a border or tint, not a figure
        if (r.width * r.height) > 0.5 * page_area:
            continue
        cx0 = max(0, int((r.x0 - pr.x0) / CELL_PT))
        cx1 = min(nx - 1, int((r.x1 - pr.x0) / CELL_PT))
        cy0 = max(0, int((r.y0 - pr.y0) / CELL_PT))
        cy1 = min(ny - 1, int((r.y1 - pr.y0) / CELL_PT))
        grid[cy0:cy1 + 1, cx0:cx1 + 1] = True
    return (grid, pr) if grid.any() else None


def _clusters_at(grid0: np.ndarray, pr: fitz.Rect, dilate: int) -> list[tuple[float, fitz.Rect]]:
    ny, nx = grid0.shape
    page_area = pr.width * pr.height
    grid = grid0
    for _ in range(dilate):
        g = grid
        grid = g.copy()
        grid[1:, :] |= g[:-1, :]
        grid[:-1, :] |= g[1:, :]
        grid[:, 1:] |= g[:, :-1]
        grid[:, :-1] |= g[:, 1:]

    seen = np.zeros_like(grid)
    rects: list[tuple[float, fitz.Rect]] = []
    for sy, sx in zip(*np.nonzero(grid)):
        sy, sx = int(sy), int(sx)
        if seen[sy, sx]:
            continue
        q = deque([(sy, sx)])
        seen[sy, sx] = True
        x0 = x1 = sx
        y0 = y1 = sy
        while q:
            y, x = q.popleft()
            x0, x1 = min(x0, x), max(x1, x)
            y0, y1 = min(y0, y), max(y1, y)
            for ny_, nx_ in ((y - 1, x), (y + 1, x), (y, x - 1), (y, x + 1)):
                if 0 <= ny_ < ny and 0 <= nx_ < nx and grid[ny_, nx_] and not seen[ny_, nx_]:
                    seen[ny_, nx_] = True
                    q.append((ny_, nx_))
        rect = fitz.Rect(
            pr.x0 + x0 * CELL_PT - PAD_PT, pr.y0 + y0 * CELL_PT - PAD_PT,
            pr.x0 + (x1 + 1) * CELL_PT + PAD_PT, pr.y0 + (y1 + 1) * CELL_PT + PAD_PT,
        ) & pr
        frac = (rect.width * rect.height) / page_area
        if MIN_AREA_FRAC <= frac <= MAX_AREA_FRAC and rect.width > 24 and rect.height > 24:
            rects.append((frac, rect))
    return rects


def _drawing_clusters(page) -> list[fitz.Rect]:
    """Group vector strokes into figure-sized rectangles.

    Dense technical sheets — a page of profile cross-sections, say — merge into
    one page-filling blob at the default dilation and would yield nothing, so we
    step the dilation down until real figures separate out.
    """
    occ = _occupancy(page)
    if occ is None:
        return []
    grid0, pr = occ
    for dilate in range(DILATE, -1, -1):
        rects = _clusters_at(grid0, pr, dilate)
        if rects:
            rects.sort(key=lambda t: -t[0])
            return [r for _, r in rects[:MAX_FIGURES_PER_PAGE]]
    return []


# --------------------------------------------------------------------------
# per-figure / per-page extraction (runs inside a worker)
# --------------------------------------------------------------------------
def _figure_row(page_num: int, kind: str, path: Path, rel: str, thumb_rel: str | None,
                w: int, h: int, area: float,
                box: tuple[float, float, float, float] = (0.0, 0.0, 0.0, 0.0)) -> tuple:
    desc = imgutil.describe(path)
    ph = dh = ed = None
    if desc:
        a, b, edge = desc
        ph, dh, ed = imgutil.hash_to_hex(a), imgutil.hash_to_hex(b), imgutil.edge_to_b64(edge)
    # `box` is the crop's position on the page, in 0-1 fractions: it lets the
    # search layer re-project page keyword hits onto the cropped drawing.
    return (page_num, kind, rel, thumb_rel, w, h, area, *box, ph, dh, ed)


def _page_fractions(rect, page_rect) -> tuple[float, float, float, float]:
    pw = page_rect.width or 1.0
    ph = page_rect.height or 1.0
    return (round((rect.x0 - page_rect.x0) / pw, 5), round((rect.y0 - page_rect.y0) / ph, 5),
            round((rect.x1 - page_rect.x0) / pw, 5), round((rect.y1 - page_rect.y0) / ph, 5))


def _save_drawings(page, out_dir: Path, page_num: int) -> list[tuple]:
    rects = _drawing_clusters(page)
    if not rects:
        return []
    zoom = FIGURE_DPI / 72
    page_area = page.rect.width * page.rect.height or 1.0
    rows = []
    for idx, rect in enumerate(rects):
        try:
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), clip=rect)
        except Exception:
            continue
        if pix.width < 40 or pix.height < 40:
            continue
        fname = f"p{page_num:03d}_fig{idx:02d}.png"
        tname = f"p{page_num:03d}_fig{idx:02d}_thumb.png"
        pix.save(out_dir / fname)
        try:
            from PIL import Image
            with Image.open(out_dir / fname) as im:
                im.thumbnail((FIG_THUMB_W, FIG_THUMB_W), Image.Resampling.LANCZOS)
                im.convert("RGB").save(out_dir / tname)
            thumb_rel = f"figures/{out_dir.name}/{tname}"
        except Exception:
            thumb_rel = f"figures/{out_dir.name}/{fname}"
        rows.append(_figure_row(page_num, "drawing", out_dir / fname,
                                f"figures/{out_dir.name}/{fname}", thumb_rel,
                                pix.width, pix.height,
                                (rect.width * rect.height) / page_area,
                                _page_fractions(rect, page.rect)))
    return rows


def _save_rasters(page, doc, out_dir: Path, page_num: int) -> list[tuple]:
    seen: set[int] = set()
    rows = []
    for idx, img in enumerate(page.get_images(full=True)):
        xref = img[0]
        if xref in seen:
            continue
        seen.add(xref)
        try:
            pix = fitz.Pixmap(doc, xref)
            if pix.colorspace is not None and pix.colorspace.n > 3:
                pix = fitz.Pixmap(fitz.csRGB, pix)
            if pix.alpha:
                pix = fitz.Pixmap(pix, 0)
            if max(pix.width, pix.height) > MAX_IMAGE_SIDE:
                factor = max(2, round(max(pix.width, pix.height) / MAX_IMAGE_SIDE))
                try:
                    pix.shrink(factor)
                except Exception:
                    pass
            if pix.width < 32 or pix.height < 32:  # spacer pixels, gradients
                continue
            fname = f"p{page_num:03d}_img{idx:02d}.png"
            pix.save(out_dir / fname)
            box = (0.0, 0.0, 0.0, 0.0)
            try:
                placed = page.get_image_rects(xref)
                if placed:
                    box = _page_fractions(placed[0], page.rect)
            except Exception:
                pass
            rows.append(_figure_row(page_num, "raster", out_dir / fname,
                                    f"figures/{out_dir.name}/{fname}", None,
                                    pix.width, pix.height, 0.0, box))
        except Exception:
            continue
    return rows


def _save_page_renders(page, pages_out: Path, page_num: int) -> None:
    page.get_pixmap(matrix=fitz.Matrix(THUMB_DPI / 72, THUMB_DPI / 72)).save(
        pages_out / f"p{page_num:03d}_thumb.png")
    page.get_pixmap(matrix=fitz.Matrix(FULL_DPI / 72, FULL_DPI / 72)).save(
        pages_out / f"p{page_num:03d}.png")


def build_pdf_payload(pdf_path_s: str, systems: list[str], known_refs: set[str],
                      summarize: bool = False) -> dict:
    """Render one PDF and return everything the parent needs to INSERT.

    Runs in a worker process: it writes PNGs (each document has its own output
    directory, so workers never collide) but touches no database.
    """
    pdf_path = Path(pdf_path_s)
    doc = fitz.open(pdf_path)
    rel = pdf_path.relative_to(PDF_DIR)
    stem = sanitize("_".join(rel.with_suffix("").parts))
    folder = rel.parts[0] if len(rel.parts) > 1 else ""

    raw_title = (doc.metadata.get("title") or "").strip()
    # several of these PDFs carry a CAD export path as their title - useless to a reader
    if not raw_title or "\\" in raw_title or raw_title.lower().endswith(".pdf"):
        raw_title = pdf_path.stem
    system = detect_system(folder, pdf_path.name)

    pages_out = PAGES_DIR / stem
    figs_out = FIGURES_DIR / stem
    pages_out.mkdir(parents=True, exist_ok=True)
    figs_out.mkdir(parents=True, exist_ok=True)

    cfg = None
    if summarize:
        import ai_client
        cfg = ai_client.load_config()

    pages, figures = [], []
    for page_num, page in enumerate(doc, start=1):
        text = page.get_text("text").strip()
        upper = text.upper()
        refs = sorted({
            r for tok in REF_RE.findall(upper)
            if (r := norm_ref(tok)) in known_refs and len(r) >= 2
            # TOC page ranges like "pages 50-51" normalize to "5051" and must not be indexed as part 5051
            and not (tok.count("-")==1 and tok.replace("-","").isdigit() and "PAGES" in upper[max(0, upper.find(tok)-30):upper.find(tok)+30])
        })
        summary = ""
        if summarize:
            try:
                import ai_client
                summary = ai_client.generate_summary(text, cfg) or ""
            except Exception:
                summary = ""
        _save_page_renders(page, pages_out, page_num)
        figures.extend(_save_rasters(page, doc, figs_out, page_num))
        figures.extend(_save_drawings(page, figs_out, page_num))
        pages.append((page_num, text, " ".join(refs), summary,
                      f"pages/{stem}/p{page_num:03d}.png"))

    n_pages = len(doc)
    doc.close()
    return {
        "document": (sha1_of(pdf_path), str(rel).replace("\\", "/"),
                     f"{system} — {raw_title}", system, json.dumps(systems),
                     " ".join(systems), detect_kind(pdf_path.name), n_pages),
        "pages": pages,
        "figures": figures,
        "label": f"{rel}: {n_pages} pages, {len(figures)} figures",
    }


def _insert_payload(conn: sqlite3.Connection, payload: dict) -> None:
    cur = conn.execute(
        "INSERT INTO documents(sha1, filename, title, system, systems, systems_text,"
        " doc_kind, num_pages) VALUES (?,?,?,?,?,?,?,?)", payload["document"])
    doc_id = cur.lastrowid
    conn.executemany(
        "INSERT INTO pages(doc_id, page_num, text, ocr_text, refs, summary)"
        " VALUES (?,?,?,?,?,?)",
        [(doc_id, num, text, "", refs, summary)
         # the 5th field is the page-render path, used by the OCR pass — it is
         # deliberately not stored: ocr_text is filled in later, or stays empty
         for num, text, refs, summary, _render in payload["pages"]])
    conn.executemany(
        "INSERT INTO figures(doc_id, page_num, kind, filename, thumb, width, height,"
        " area, x0, y0, x1, y1, phash, dhash, edge)"
        " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [(doc_id, *row) for row in payload["figures"]])


# --------------------------------------------------------------------------
# workbook + reference photos
# --------------------------------------------------------------------------
def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def index_xlsx(conn: sqlite3.Connection) -> int:
    """Index the PROFILES & ACCESSORIES workbook by column, not by position.

    PROFILES is `Ref.No. | Description`; ACCESSORIES is
    `Item No. | Supplier No. | (blank) | Designation`. The header row is located
    first so the leading junk rows are skipped.
    """
    import openpyxl

    files = sorted(p for p in PDF_DIR.rglob("*")
                   if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"}
                   and not p.name.startswith("~$"))
    count = 0
    for x in files:
        try:
            wb = openpyxl.load_workbook(x, read_only=True, data_only=True)
        except Exception as exc:
            print(f"  ! cannot read {x.name}: {exc}", flush=True)
            continue
        for sheet in wb.sheetnames:
            name_l = sheet.lower()
            kind = "profile" if "profile" in name_l else (
                "accessory" if "accessor" in name_l else "part")
            ws = wb[sheet]
            col_ref = col_sup = col_des = None
            for row in ws.iter_rows(values_only=True):
                cells = [_cell(v) for v in row]
                if col_ref is None:
                    low = [c.lower() for c in cells]
                    for i, c in enumerate(low):
                        if "ref" in c or "item" in c:
                            col_ref = i
                        elif "supplier" in c:
                            col_sup = i
                        elif "design" in c or "descri" in c:
                            col_des = i
                    if col_ref is None:
                        continue          # still above the header row
                    if col_des is None:   # PROFILES: description is the next filled column
                        col_des = col_ref + 1
                    continue              # skip the header row itself

                ref = cells[col_ref] if col_ref < len(cells) else ""
                if not ref:
                    continue
                ref = ref.upper()
                designation = cells[col_des] if col_des is not None and col_des < len(cells) else ""
                supplier = cells[col_sup] if col_sup is not None and col_sup < len(cells) else ""
                # anything else on the row still belongs in the searchable blob
                extra = [c for i, c in enumerate(cells)
                         if c and i not in {col_ref, col_sup, col_des}]
                if not designation and extra:
                    designation = " - ".join(extra)
                    extra = []
                raw = " ".join(filter(None, [ref, kind, designation, supplier, *extra]))
                conn.execute(
                    "INSERT INTO parts(ref, ref_norm, kind, supplier_ref, designation, raw)"
                    " VALUES (?,?,?,?,?,?)",
                    (ref, norm_ref(ref), kind, supplier, designation, raw),
                )
                count += 1
        wb.close()
        print(f"  {x.name}: {count} rows", flush=True)
    return count


def hash_ref_chunk(names: list[str]) -> list[tuple]:
    """Descriptors for a batch of part photos (runs in a worker process)."""
    rows = []
    for name in names:
        p = REF_IMAGES_DIR / name
        ref = Path(name).stem.strip().upper()
        desc = imgutil.describe(p)
        ph = dh = ed = None
        if desc:
            a, b, edge = desc
            ph, dh, ed = imgutil.hash_to_hex(a), imgutil.hash_to_hex(b), imgutil.edge_to_b64(edge)
        rows.append((ref, norm_ref(ref), name, ph, dh, ed))
    return rows


def index_reference_images(conn: sqlite3.Connection, workers: int) -> int:
    """Hash every part photo in pdfs/Images; the filename stem is the part ref."""
    if not REF_IMAGES_DIR.is_dir():
        return 0
    exts = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}
    names = sorted(p.relative_to(REF_IMAGES_DIR).as_posix()
                   for p in REF_IMAGES_DIR.rglob("*")
                   if p.is_file() and p.suffix.lower() in exts)
    if not names:
        return 0
    chunks = [names[i:i + REF_CHUNK] for i in range(0, len(names), REF_CHUNK)]
    done = 0

    def store(rows):
        nonlocal done
        conn.executemany(
            "INSERT INTO ref_images(ref, ref_norm, filename, phash, dhash, edge)"
            " VALUES (?,?,?,?,?,?)", rows)
        done += len(rows)
        print(f"  hashed {done}/{len(names)} part photos…", flush=True)

    if workers > 1 and len(chunks) > 1:
        try:
            with ProcessPoolExecutor(max_workers=workers) as pool:
                for fut in as_completed([pool.submit(hash_ref_chunk, c) for c in chunks]):
                    store(fut.result())
            return done
        except Exception as exc:
            print(f"  ! parallel hashing unavailable ({exc}); falling back to one core",
                  flush=True)
            conn.execute("DELETE FROM ref_images")
            done = 0
    for c in chunks:
        store(hash_ref_chunk(c))
    return done


# --------------------------------------------------------------------------
# OCR + vectors
# --------------------------------------------------------------------------
def run_ocr(conn: sqlite3.Connection, mode: str = "auto") -> int:
    """Read text off the page renders where the PDF text layer fails us.

    `auto` only touches pages that are sparse or damaged (about one in five here);
    `all` transcribes every page, which is slower and, on a hosted model, costs
    money — worth it when you want maximum recall.
    """
    import ai_client
    import ocr as ocr_mod
    import search as search_mod

    if mode == "off":
        return 0
    cfg = ai_client.load_config()
    engine = ocr_mod.pick_engine(cfg, mode if mode in ("vision", "easyocr") else "auto")
    if not engine:
        print("  (no OCR engine available - skipping)", flush=True)
        return 0

    rows = conn.execute(
        "SELECT p.id, p.text, d.filename, p.page_num FROM pages p "
        "JOIN documents d ON d.id = p.doc_id ORDER BY p.id").fetchall()
    todo = []
    for pid, text, filename, page_num in rows:
        stem = sanitize("_".join(Path(filename).with_suffix("").parts))
        image = PAGES_DIR / stem / f"p{page_num:03d}.png"
        if mode == "all" or ocr_mod.needs_ocr(text or "", search_mod.clean_text(text or "")):
            todo.append((pid, image))
    if not todo:
        return 0

    print(f"OCR: {len(todo)} page(s) via {engine}...", flush=True)
    done = 0
    for pid, image in todo:
        text = ocr_mod.transcribe(image, cfg, engine)
        if text:
            conn.execute("UPDATE pages SET ocr_text = ? WHERE id = ?", (text, pid))
            done += 1
        # geometry for highlighting: these are exactly the pages whose text layer
        # PyMuPDF cannot search, so without boxes they could never be highlighted
        meta = conn.execute("SELECT doc_id, page_num FROM pages WHERE id = ?",
                            (pid,)).fetchone()
        if meta:
            boxes = ocr_mod.word_boxes(image)
            if boxes:
                conn.execute("DELETE FROM ocr_boxes WHERE doc_id = ? AND page_num = ?",
                             (meta[0], meta[1]))
                conn.executemany(
                    "INSERT INTO ocr_boxes(doc_id, page_num, text, x, y, w, h)"
                    " VALUES (?,?,?,?,?,?,?)",
                    [(meta[0], meta[1], t, *b) for t, b in boxes])
        if done and done % 10 == 0:
            conn.commit()
            print(f"  OCR {done}/{len(todo)}...", flush=True)
    conn.commit()
    print(f"  OCR recovered text on {done} page(s)", flush=True)
    return done


def build_vectors(conn: sqlite3.Connection) -> int:
    """Embed every page and part so semantic queries can find them."""
    import embed
    import search as search_mod

    if not embed.available():
        print("  (sentence-transformers not installed - skipping vectors)", flush=True)
        return 0

    pages = conn.execute(
        "SELECT p.id, p.text, p.ocr_text, d.title, d.systems_text, d.doc_kind "
        "FROM pages p JOIN documents d ON d.id = p.doc_id ORDER BY p.id").fetchall()
    page_texts = [
        f"{r[3]} | {r[4]} | {r[5]}\n{search_mod.clean_text(r[1] or '')}\n{r[2] or ''}"[:2000]
        for r in pages
    ]
    print(f"Embedding {len(page_texts)} pages...", flush=True)
    vecs = embed.encode(page_texts, progress=False)
    if vecs is None:
        print("  (embedding model unavailable - skipping vectors)", flush=True)
        return 0
    conn.executemany("UPDATE pages SET vec = ? WHERE id = ?",
                     [(embed.to_b64(v), r[0]) for v, r in zip(vecs, pages)])

    parts = conn.execute(
        "SELECT id, ref, kind, designation FROM parts ORDER BY id").fetchall()
    part_texts = [f"{r[1]} {r[2]} {r[3]}".strip() for r in parts]
    print(f"Embedding {len(part_texts)} parts...", flush=True)
    pvecs = embed.encode(part_texts, batch_size=256, progress=False)
    if pvecs is not None:
        conn.executemany("UPDATE parts SET vec = ? WHERE id = ?",
                         [(embed.to_b64(v), r[0]) for v, r in zip(pvecs, parts)])
    conn.commit()
    total = len(page_texts) + (len(part_texts) if pvecs is not None else 0)
    print(f"  {total} vectors ({embed.model_name()})", flush=True)
    return total


# --------------------------------------------------------------------------
# driver
# --------------------------------------------------------------------------
def build_visual_index(conn: sqlite3.Connection) -> int:
    """Patch-embed every figure and part photo so a *crop* can find its source.

    Whole-image hashes cannot do this: cropping changes the hash completely. See
    vismatch for the recall/verification split.
    """
    try:
        import vismatch
    except Exception as exc:
        print(f"  (visual index unavailable: {str(exc)[:120]})", flush=True)
        return 0
    if not vismatch.available():
        print("  (visual encoder unavailable - skipping partial-image search)", flush=True)
        return 0

    items = []
    for r in conn.execute(
            "SELECT f.filename, f.page_num, d.filename AS doc, d.system "
            "FROM figures f JOIN documents d ON d.id = f.doc_id").fetchall():
        items.append({"key": f"fig:{r[0]}", "kind": "catalogue", "file": r[0],
                      "page_num": r[1], "doc": r[2], "system": r[3],
                      "path": str(DATA_DIR / r[0])})
    for r in conn.execute("SELECT ref, filename FROM ref_images").fetchall():
        items.append({"key": f"part:{r[1]}", "kind": "part", "file": r[1], "ref": r[0],
                      "path": str(REF_IMAGES_DIR / r[1])})
    print(f"Visual index: embedding {len(items)} images "
          f"({vismatch.TILES_PER_IMAGE} regions each)...", flush=True)
    return vismatch.build_index(items,
                                progress=lambda m: print(m, flush=True))


def rebuild(summarize: bool = False, workers: int = 0,
            ocr_mode: str = "auto") -> int:
    """Wipe and rebuild the whole index. Returns the number of documents indexed."""
    import time
    t0 = time.time()
    n_workers = worker_count(workers)

    cache = DATA_DIR / "cache.json"
    saved_cache = cache.read_bytes() if cache.exists() else None
    for sub in (PAGES_DIR, FIGURES_DIR):
        shutil.rmtree(sub, ignore_errors=True)
    shutil.rmtree(DATA_DIR / "images", ignore_errors=True)  # pre-1.1 layout
    for leftover in DATA_DIR.glob("index.db*"):
        leftover.unlink(missing_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if saved_cache is not None:
        cache.write_bytes(saved_cache)  # answers survive a re-index

    conn = sqlite3.connect(DB_PATH, timeout=60)
    init_db(conn)

    # workbook first: page-level reference detection validates against it
    print("Indexing workbook…", flush=True)
    n_parts = index_xlsx(conn)
    conn.execute("INSERT INTO parts_fts(rowid, ref, designation) "
                 "SELECT id, ref, designation FROM parts")
    print(f"Hashing part photos on {n_workers} core(s)…", flush=True)
    n_refs = index_reference_images(conn, n_workers)
    conn.commit()

    known_refs = {r[0] for r in conn.execute("SELECT ref_norm FROM parts").fetchall()}
    known_refs |= {r[0] for r in conn.execute("SELECT ref_norm FROM ref_images").fetchall()}
    known_refs.discard("")

    online = False
    if summarize:
        import ai_client
        online, _ = ai_client.check_online()
        if not online:
            print("  (AI offline — skipping page summaries)", flush=True)
    do_summary = summarize and online

    pdfs = sorted(p for p in PDF_DIR.rglob("*") if p.is_file() and p.suffix.lower() == ".pdf")
    # each PDF becomes its own document (may share content with other folders)
    jobs = []
    seen_stems: set[str] = set()
    for p in pdfs:
        stem = p.stem.lower()
        # Allow duplicate stems if they're in different folders (e.g., Comete vs Galaxie)
        systems = sorted({
            detect_system(p.relative_to(PDF_DIR).parts[0]
                          if len(p.relative_to(PDF_DIR).parts) > 1 else "", p.name)
            for p in [p]
        })
        jobs.append((str(p), systems))

    print(f"Rendering {len(jobs)} unique PDF(s) from {len(pdfs)} file(s) "
          f"on {n_workers} core(s)…", flush=True)
    payloads = []
    if n_workers > 1 and len(jobs) > 1 and not do_summary:
        try:
            with ProcessPoolExecutor(max_workers=min(n_workers, len(jobs))) as pool:
                futures = [pool.submit(build_pdf_payload, path, systems, known_refs, False)
                           for path, systems in jobs]
                for fut in as_completed(futures):
                    payloads.append(fut.result())
        except Exception as exc:
            print(f"  ! parallel rendering unavailable ({exc}); falling back to one core",
                  flush=True)
            payloads = []
    if not payloads:
        payloads = [build_pdf_payload(path, systems, known_refs, do_summary)
                    for path, systems in jobs]

    for payload in sorted(payloads, key=lambda p: p["document"][1]):
        _insert_payload(conn, payload)
        print(f"  {payload['label']}", flush=True)

    n_ocr = run_ocr(conn, ocr_mode)
    n_vec = build_vectors(conn)
    n_vis = build_visual_index(conn)

    conn.execute(
        "INSERT INTO pages_fts(rowid, text, refs, title, ocr) "
        "SELECT p.id, p.text || ' ' || COALESCE(p.summary,''), p.refs, "
        "       d.title || ' ' || d.systems_text || ' ' || d.doc_kind, "
        "       COALESCE(p.ocr_text,'') "
        "FROM pages p JOIN documents d ON d.id = p.doc_id"
    )
    conn.execute("INSERT INTO pages_fts(pages_fts) VALUES('optimize')")
    conn.execute("INSERT INTO parts_fts(parts_fts) VALUES('optimize')")
    conn.commit()

    n_pages = conn.execute("SELECT COUNT(*) FROM pages").fetchone()[0]
    n_figs = conn.execute("SELECT COUNT(*) FROM figures").fetchone()[0]
    conn.close()
    print(f"Done in {time.time() - t0:.0f}s: {len(payloads)} documents, {n_pages} pages, "
          f"{n_figs} figures, {n_parts} parts, {n_refs} part photos, "
          f"{n_ocr} pages OCR'd, {n_vec} vectors, "
          f"{n_vis} images visually indexed.", flush=True)
    return len(payloads)


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Rebuild the Installux search index.")
    ap.add_argument("--summarize", action="store_true",
                    help="also ask the local LLM for a one-line summary per page (slow, serial)")
    ap.add_argument("--workers", type=int, default=0,
                    help="CPU worker processes (default: one per core, max 8)")
    ap.add_argument("--ocr", default="auto", choices=["auto", "all", "vision", "easyocr", "off"],
                    help="auto: only sparse/damaged pages (default); all: every page")
    args = ap.parse_args()
    rebuild(summarize=args.summarize, workers=args.workers, ocr_mode=args.ocr)
